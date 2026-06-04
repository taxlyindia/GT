# routers/customers.py — Customer CRUD (mobile = primary key per tenant)

from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import pandas as pd

from database import get_db
from models import Customer, Invoice, Payment, Advance
from utils.auth import get_tenant_payload as get_current_user_payload
from utils.business import pan_is_mandatory, is_sft_flagged, current_fy

router = APIRouter()


# ── Schemas ───────────────────────────────────────────────────

class CustomerCreate(BaseModel):
    mobile:  str   = Field(..., pattern=r"^\d{10}$")   # PRIMARY KEY
    name:    str   = Field(..., min_length=1, max_length=200)
    state:   str   = Field(..., min_length=2)            # Mandatory for GST
    pan:     Optional[str] = Field(None, pattern=r"^[A-Z]{5}[0-9]{4}[A-Z]$")
    gstin:   Optional[str] = None
    address: Optional[str] = None
    email:   Optional[str] = None

class CustomerOut(BaseModel):
    mobile:           str
    name:             str
    pan:              Optional[str]
    state:            str
    gstin:            Optional[str]
    address:          Optional[str]
    cash_receipts_fy: float
    sft_flagged:      bool
    pan_mandatory:    bool

    class Config:
        from_attributes = True

    @classmethod
    def from_orm_with_extras(cls, c: Customer):
        return cls(
            mobile=c.mobile,
            name=c.name,
            pan=c.pan,
            state=c.state,
            gstin=c.gstin,
            address=c.address,
            cash_receipts_fy=float(c.cash_receipts_fy),
            sft_flagged=c.sft_flagged,
            pan_mandatory=pan_is_mandatory(c.cash_receipts_fy),
        )


# ── Create / Update Customer ──────────────────────────────────

@router.post("/", response_model=CustomerOut, status_code=201)
async def create_customer(
    body:    CustomerCreate,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """Create a new customer. Mobile is the unique identifier per tenant."""
    tenant_id = payload["tenant_id"]
    existing = await db.get(Customer, (body.mobile, tenant_id))
    if existing:
        raise HTTPException(status_code=409, detail="Customer with this mobile already exists.")

    customer = Customer(
        mobile=body.mobile,
        tenant_id=tenant_id,
        name=body.name,
        pan=body.pan,
        state=body.state,
        gstin=body.gstin,
        address=body.address,
        email=body.email,
        cash_receipts_fy=0,
        sft_flagged=False,
    )
    db.add(customer)
    await db.commit()
    await db.refresh(customer)
    return CustomerOut.from_orm_with_extras(customer)


@router.put("/{mobile}", response_model=CustomerOut)
async def update_customer(
    mobile:  str,
    body:    CustomerCreate,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    tenant_id = payload["tenant_id"]
    customer  = await db.get(Customer, (mobile, tenant_id))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    for field, value in body.dict(exclude_unset=True, exclude={"mobile"}).items():
        setattr(customer, field, value)

    await db.commit()
    await db.refresh(customer)
    return CustomerOut.from_orm_with_extras(customer)


# ── List Customers ────────────────────────────────────────────

@router.get("/", response_model=list[CustomerOut])
async def list_customers(
    q:       Optional[str] = None,
    limit:   Optional[int] = None,
    payload: dict          = Depends(get_current_user_payload),
    db:      AsyncSession  = Depends(get_db),
):
    """List customers with optional name/mobile search."""
    tenant_id = payload["tenant_id"]
    stmt = select(Customer).where(Customer.tenant_id == tenant_id).order_by(Customer.name)

    if q:
        stmt = stmt.where(
            Customer.name.ilike(f"%{q}%") | Customer.mobile.contains(q) |
            Customer.pan.ilike(f"%{q}%")
        )
    if limit:
        stmt = stmt.limit(limit)

    result = await db.execute(stmt)
    customers = result.scalars().all()
    return [CustomerOut.from_orm_with_extras(c) for c in customers]


# ── Customer Ledger ───────────────────────────────────────────

@router.get("/{mobile}/ledger")
async def customer_ledger(
    mobile:  str,
    payload: dict         = Depends(get_current_user_payload),
    db:      AsyncSession = Depends(get_db),
):
    """
    Return full transaction ledger for a customer:
    Invoices (debit) + Payments (credit) + Advances, sorted by date.
    """
    tenant_id = payload["tenant_id"]
    customer  = await db.get(Customer, (mobile, tenant_id))
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Invoices
    inv_result = await db.execute(
        select(Invoice).where(
            Invoice.tenant_id      == tenant_id,
            Invoice.customer_mobile == mobile,
        ).order_by(Invoice.invoice_date)
    )
    invoices = inv_result.scalars().all()

    # Payments
    pay_result = await db.execute(
        select(Payment).where(
            Payment.tenant_id      == tenant_id,
            Payment.customer_mobile == mobile,
        ).order_by(Payment.payment_date)
    )
    payments = pay_result.scalars().all()

    # Advances (full amount when recorded)
    adv_result = await db.execute(
        select(Advance).where(
            Advance.tenant_id      == tenant_id,
            Advance.customer_mobile == mobile,
        ).order_by(Advance.advance_date)
    )
    advances = adv_result.scalars().all()

    # Build ledger entries — unsorted, then sort by date
    raw_entries = []

    for inv in invoices:
        raw_entries.append({
            "date":   inv.invoice_date.isoformat(),
            "type":   "Invoice",
            "ref":    inv.invoice_no,
            "debit":  float(inv.grand_total),
            "credit": 0.0,
        })

    for pay in payments:
        raw_entries.append({
            "date":   pay.payment_date.isoformat(),
            "type":   "Payment",
            "ref":    f"PMT-{pay.id}",
            "debit":  0.0,
            "credit": float(pay.amount),
        })

    for adv in advances:
        # Show the advance as a credit entry (cash received from customer).
        # AdvanceAllocation rows are intentionally excluded — allocation is an internal
        # application of the advance against an invoice, not a new cash receipt.
        # Showing both would double-count the same payment.
        raw_entries.append({
            "date":   adv.advance_date.isoformat(),
            "type":   "Advance",
            "ref":    f"ADV-{adv.id}",
            "debit":  0.0,
            "credit": float(adv.amount),
        })

    raw_entries.sort(key=lambda e: e["date"])

    # Running balance
    entries = []
    balance = 0.0
    for e in raw_entries:
        balance += e["debit"] - e["credit"]
        entries.append({**e, "balance": round(balance, 2)})

    return {
        "customer":        {"mobile": customer.mobile, "name": customer.name, "pan": customer.pan},
        "outstanding":     round(balance, 2),
        "total_invoiced":  sum(e["debit"]  for e in entries),
        "total_paid":      sum(e["credit"] for e in entries),
        "entries":         entries,
    }


# ── Bulk Import via Excel ─────────────────────────────────────

@router.get("/excel-template")
async def download_customer_template(
    _: dict = Depends(get_current_user_payload),
):
    """
    Download a blank Excel template for bulk customer import.
    Columns: mobile, name, state, pan, gstin, address
    Fixed per Improvement document request.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Import"

    headers = ["mobile", "name", "state", "pan", "gstin", "address"]
    ws.append(headers)

    gold_fill   = PatternFill("solid", fgColor="C8900A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = gold_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Sample row
    ws.append(["9876543210", "Ramesh Kumar", "Maharashtra", "ABCDE1234F", "", "123 Main St"])

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required", "Notes"])
    ws2.append(["mobile", "YES", "10-digit Indian mobile number"])
    ws2.append(["name", "YES", "Customer full name"])
    ws2.append(["state", "YES", "Indian state name"])
    ws2.append(["pan", "No", "PAN card (ABCDE1234F format). Mandatory if cash FY > ₹2L or invoice > ₹2L."])
    ws2.append(["gstin", "No", "GST number if registered"])
    ws2.append(["address", "No", "Customer address"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customer_import_template.xlsx"'},
    )


REQUIRED_COLS = {"mobile", "name", "state"}

@router.post("/import-excel")
async def import_customers_excel(
    file:    UploadFile       = File(...),
    payload: dict             = Depends(get_current_user_payload),
    db:      AsyncSession     = Depends(get_db),
):
    """
    Bulk-import customers from an Excel file.
    Required columns: mobile, name, state
    Optional columns: pan, gstin, address, email
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx or .xls)")

    contents = await file.read()
    df = pd.read_excel(BytesIO(contents), dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]

    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(missing)}. "
                   f"Download the template from Customer Master → Import Excel."
        )

    tenant_id = payload["tenant_id"]
    created = 0
    updated = 0
    errors  = []

    for i, row in df.iterrows():
        mobile = str(row.get("mobile", "")).strip()
        name   = str(row.get("name",   "")).strip()
        state  = str(row.get("state",  "")).strip()

        if not mobile or not name or not state:
            errors.append(f"Row {i+2}: mobile, name, state are required")
            continue
        if len(mobile) != 10 or not mobile.isdigit():
            errors.append(f"Row {i+2}: invalid mobile '{mobile}'")
            continue

        existing = await db.get(Customer, (mobile, tenant_id))
        if existing:
            existing.name    = name
            existing.state   = state
            existing.pan     = str(row.get("pan", "")).strip() or existing.pan
            existing.gstin   = str(row.get("gstin", "")).strip() or existing.gstin
            existing.address = str(row.get("address", "")).strip() or existing.address
            updated += 1
        else:
            db.add(Customer(
                mobile=mobile, tenant_id=tenant_id,
                name=name, state=state,
                pan=str(row.get("pan", "")).strip() or None,
                gstin=str(row.get("gstin", "")).strip() or None,
                address=str(row.get("address", "")).strip() or None,
            ))
            created += 1

    await db.commit()
    return {"created": created, "updated": updated, "errors": errors}
